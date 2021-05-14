from openpyxl.styles import NamedStyle, Font, Color, Alignment, Border, Side, PatternFill, colors

# Weisser Hintergrund
format_background = NamedStyle(name="format_background")
format_background.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# Format einzelne Races und Saison Rangliste
# Die zwei Titelzeilen
format_title = NamedStyle(name="format_title")
format_title.font = Font(size=20,
                  bold=True,
                  color='2E373D')

format_background.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# Listenheader zentriert
format_header_center = NamedStyle(name="format_header_center")
format_header_center.font = Font(size=12,
                   bold=True,
                   color='FFFFFF')

format_header_center.alignment = Alignment(horizontal="center",
                             vertical="center")

format_header_center.fill = PatternFill(start_color="2E373D",
                          end_color="2E373D",
                          fill_type='solid')


# Listenheader links
format_header_left = NamedStyle(name="format_header_left")
format_header_left.font = Font(size=12,
                   bold=True,
                   color='FFFFFF')

format_header_left.alignment = Alignment(horizontal="left",
                             vertical="center")

format_header_left.fill = PatternFill(start_color="2E373D",
                          end_color="2E373D",
                          fill_type='solid')


# Spalte Position
format_pos = NamedStyle(name="format_pos")
format_pos.font = Font(size=18, color='000000', bold=True)

format_pos.border = Border(bottom=Side(border_style='thin'))

format_pos.alignment = Alignment(horizontal="center",
                             vertical="center")

format_pos.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# Spalte Driver
format_driver = NamedStyle(name="format_driver")
format_driver.font = Font(size=16, color='000000', bold=True)

format_driver.border = Border(bottom=Side(border_style='thin'))

format_driver.alignment = Alignment(horizontal="left",
                             vertical="center")

format_driver.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# Spalte Team
format_team = NamedStyle(name="format_team")
format_team.font = Font(size=16, color='000000')

format_team.border = Border(bottom=Side(border_style='thin'))

format_team.alignment = Alignment(horizontal="left",
                             vertical="center")

format_team.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# Spalte Starts und Incs
format_start_inc = NamedStyle(name="format_start_inc")
format_start_inc.font = Font(size=16, color='000000')

format_start_inc.border = Border(bottom=Side(border_style='thin'))

format_start_inc.alignment = Alignment(horizontal="center",
                             vertical="center")

format_start_inc.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# Spalte Points
format_points = NamedStyle(name="format_points")
format_points.font = Font(size=16, bold=True, color='000000')

format_points.border = Border(bottom=Side(border_style='thin'))

format_points.alignment = Alignment(horizontal="center",
                             vertical="center")

format_points.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# Sorterhinweis
format_sort_hint = NamedStyle(name="format_sort_hint")

format_sort_hint.font = Font(size=12, color='000000')

format_sort_hint.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')