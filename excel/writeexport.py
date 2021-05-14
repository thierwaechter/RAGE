from openpyxl import Workbook
import pandas as pd
from excel.formatexport import (
                                format_background,
                                format_title,
                                format_driver,
                                format_header_center,
                                format_header_left,
                                format_points,
                                format_pos,
                                format_start_inc,
                                format_team,
                                format_sort_hint)

workbook = Workbook()
dest_filename = 'RAGE Export.xlsx'
workbook.save(filename=dest_filename)

book = Workbook()
#book = load_workbook('RAGE Export.xlsx')
writer = pd.ExcelWriter('RAGE Export.xlsx', engine='openpyxl', mode='a')
writer.book = book

def format_team_export(payload):
    dataframe = payload[0]
    date = payload[1]
    saison = payload[2]
    sheetname_long = "Team " + saison
    sheetname = sheetname_long[:30]
    if 'Sheet' in writer.book.sheetnames:
        del writer.book['Sheet']
    dataframe.to_excel(writer, sheet_name=sheetname, startrow=3, index=False)
    writer.book[sheetname].column_dimensions['A'].width = 10
    writer.book[sheetname].column_dimensions['B'].width = 90
    writer.book[sheetname].column_dimensions['C'].width = 10
    writer.book[sheetname].column_dimensions['D'].width = 10
    writer.book[sheetname].column_dimensions['E'].width = 10
    # Formatiere Datum
    writer.book[sheetname].cell(row=1, column=1, value="Erstellt am " + date).style = format_title
    # Formatiere Streckenname
    writer.book[sheetname].cell(row=2, column=1, value="Für die Saison: " + saison).style = format_title
    # Formatiere Header
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=1, max_col=5, max_row=4):
        for cell in row:
            cell.style = format_header_center
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=2, max_col=3, max_row=4):
        for cell in row:
            cell.style = format_header_left
    # Suche letzte Zeile
    row_line = 3
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=2, max_col=2, max_row=4000):
        for cell in row:
            if not cell.value is None:
                row_line += 1
    # Formatiere Position
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=1, max_col=1, max_row=row_line):
        for cell in row:
            cell.style = format_pos
    # Formatiere Driver
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=2, max_col=2, max_row=row_line):
        for cell in row:
            cell.style = format_driver
    # Formatiere Team
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=3, max_col=3, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Formatiere Starts und Incs
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=3, max_col=4, max_row=row_line):
        for cell in row:
            cell.style = format_start_inc
    # Formatiere Points
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=5, max_col=5, max_row=row_line):
        for cell in row:
            cell.style = format_points
    # Füge Sortier Erkärung hinzu
    writer.book[sheetname].cell(
                                row=row_line + 1, 
                                column=1, 
                                value="Reihenfolge = A) Punkte, B) Anzahl Starts, C) Wenigste Inc.").style = format_sort_hint
    # Weisser Hintergrund
    for row in writer.book[sheetname].iter_rows(min_row=1, min_col=2, max_col=5, max_row=2):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=3, min_col=1, max_col=5, max_row=3):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=row_line + 1, min_col=2, max_col=5, max_row=row_line + 1):
        for cell in row:
            cell.style = format_background
    writer.save()



def format_race_export(payload):
    dataframe = payload[0]
    date = payload[1]
    saison = payload[2]
    sheetname_long = "Fahrer " + saison
    sheetname = sheetname_long[:30]
    if 'Sheet' in writer.book.sheetnames:
        del writer.book['Sheet']
    dataframe.to_excel(writer, sheet_name=sheetname, startrow=3, index=False)
    writer.book[sheetname].column_dimensions['A'].width = 10
    writer.book[sheetname].column_dimensions['B'].width = 40
    writer.book[sheetname].column_dimensions['C'].width = 50
    writer.book[sheetname].column_dimensions['D'].width = 10
    writer.book[sheetname].column_dimensions['E'].width = 10
    writer.book[sheetname].column_dimensions['F'].width = 10
    # Formatiere Datum
    writer.book[sheetname].cell(row=1, column=1, value="Erstellt am " + date).style = format_title
    # Formatiere Streckenname
    writer.book[sheetname].cell(row=2, column=1, value="Für die Saison: " + saison).style = format_title
    # Formatiere Header
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=1, max_col=6, max_row=4):
        for cell in row:
            cell.style = format_header_center
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=2, max_col=3, max_row=4):
        for cell in row:
            cell.style = format_header_left
    # Suche letzte Zeile
    row_line = 3
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=2, max_col=2, max_row=4000):
        for cell in row:
            if not cell.value is None:
                row_line += 1
    # Formatiere Position
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=1, max_col=1, max_row=row_line):
        for cell in row:
            cell.style = format_pos
    # Formatiere Driver
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=2, max_col=2, max_row=row_line):
        for cell in row:
            cell.style = format_driver
    # Formatiere Team
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=3, max_col=3, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Formatiere Starts und Incs
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=4, max_col=5, max_row=row_line):
        for cell in row:
            cell.style = format_start_inc
    # Formatiere Points
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=6, max_col=6, max_row=row_line):
        for cell in row:
            cell.style = format_points
    writer.book[sheetname].cell(
                                row=row_line + 1, 
                                column=1, 
                                value="Reihenfolge = A) Punkte, B) Anzahl Starts, C) Wenigste Inc.").style = format_sort_hint
    # Weisser Hintergrund
    for row in writer.book[sheetname].iter_rows(min_row=1, min_col=2, max_col=6, max_row=2):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=3, min_col=1, max_col=6, max_row=3):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=row_line + 1, min_col=2, max_col=6, max_row=row_line + 1):
        for cell in row:
            cell.style = format_background
    writer.save()

def format_all_race_export(payload):
    dataframe = payload[0]
    date = payload[1]
    serie = payload[2]
    sheetname_long = "Alle Rennen " + date
    sheetname = sheetname_long[:30]
    if 'Sheet' in writer.book.sheetnames:
        del writer.book['Sheet']
    dataframe.to_excel(writer, sheet_name=sheetname, startrow=3, index=False)
    writer.book[sheetname].column_dimensions['A'].width = 60
    writer.book[sheetname].column_dimensions['B'].width = 20
    writer.book[sheetname].column_dimensions['C'].width = 10
    writer.book[sheetname].column_dimensions['D'].width = 40
    writer.book[sheetname].column_dimensions['E'].width = 50
    writer.book[sheetname].column_dimensions['F'].width = 10
    writer.book[sheetname].column_dimensions['G'].width = 10
    writer.book[sheetname].column_dimensions['H'].width = 10
    # Formatiere Datum
    writer.book[sheetname].cell(row=1, column=1, value="Erstellt am " + date).style = format_title
    # Formatiere Streckenname
    writer.book[sheetname].cell(row=2, column=1, value="Für die Saison: " + serie).style = format_title
    # Formatiere Header linksbündig für Track und Datum
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=1, max_col=2, max_row=4):
        for cell in row:
            cell.style = format_header_left
    # Formatiere Header entriert für Pos
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=3, max_col=3, max_row=4):
        for cell in row:
            cell.style = format_header_center
    # Formatiere Header linksbündig für Driver und Team
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=4, max_col=5, max_row=4):
        for cell in row:
            cell.style = format_header_left
    # Formatiere Header entriert für Starts, Incs und Points
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=6, max_col=8, max_row=4):
        for cell in row:
            cell.style = format_header_center
    # Suche letzte Zeile
    row_line = 3
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=2, max_col=2, max_row=4000):
        for cell in row:
            if not cell.value is None:
                row_line += 1
    # Formatiere Track und Datum
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=1, max_col=2, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Formatiere Position
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=3, max_col=3, max_row=row_line):
        for cell in row:
            cell.style = format_pos
    # Formatiere Driver
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=4, max_col=4, max_row=row_line):
        for cell in row:
            cell.style = format_driver
    # Formatiere Team
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=5, max_col=5, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Formatiere Starts und Incs
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=6, max_col=7, max_row=row_line):
        for cell in row:
            cell.style = format_start_inc
    # Formatiere Points
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=8, max_col=8, max_row=row_line):
        for cell in row:
            cell.style = format_points
    # Weisser Hintergrund
    for row in writer.book[sheetname].iter_rows(min_row=1, min_col=2, max_col=8, max_row=2):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=3, min_col=1, max_col=8, max_row=3):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=row_line + 1, min_col=2, max_col=8, max_row=row_line + 1):
        for cell in row:
            cell.style = format_background
    writer.save()


def format_all_drivers_export(payload):
    dataframe = payload[0]
    date = payload[1]
    name = payload[2]
    sheetname_long = "Alle Fahrer " + date
    sheetname = sheetname_long[:30]
    if 'Sheet' in writer.book.sheetnames:
        del writer.book['Sheet']
    dataframe.to_excel(writer, sheet_name=sheetname, startrow=3, index=False)
    writer.book[sheetname].column_dimensions['A'].width = 20
    writer.book[sheetname].column_dimensions['B'].width = 40
    writer.book[sheetname].column_dimensions['C'].width = 30
    writer.book[sheetname].column_dimensions['D'].width = 50
    writer.book[sheetname].column_dimensions['E'].width = 20
    writer.book[sheetname].column_dimensions['F'].width = 50
    writer.book[sheetname].column_dimensions['G'].width = 10
    writer.book[sheetname].column_dimensions['H'].width = 10
    writer.book[sheetname].column_dimensions['I'].width = 10
    writer.book[sheetname].column_dimensions['J'].width = 20
    writer.book[sheetname].column_dimensions['K'].width = 50
    writer.book[sheetname].column_dimensions['L'].width = 20
    writer.book[sheetname].column_dimensions['M'].width = 50
    writer.book[sheetname].column_dimensions['N'].width = 10
    writer.book[sheetname].column_dimensions['O'].width = 10
    writer.book[sheetname].column_dimensions['P'].width = 10
    writer.book[sheetname].column_dimensions['Q'].width = 20
    writer.book[sheetname].column_dimensions['R'].width = 50
    writer.book[sheetname].column_dimensions['S'].width = 20
    writer.book[sheetname].column_dimensions['T'].width = 50
    writer.book[sheetname].column_dimensions['U'].width = 10
    writer.book[sheetname].column_dimensions['V'].width = 10
    writer.book[sheetname].column_dimensions['W'].width = 10
    writer.book[sheetname].column_dimensions['X'].width = 20
    writer.book[sheetname].column_dimensions['Y'].width = 50
    writer.book[sheetname].column_dimensions['Z'].width = 20
    writer.book[sheetname].column_dimensions['AA'].width = 50
    writer.book[sheetname].column_dimensions['AB'].width = 10
    writer.book[sheetname].column_dimensions['AC'].width = 10
    writer.book[sheetname].column_dimensions['AD'].width = 10
    writer.book[sheetname].column_dimensions['AE'].width = 20
    writer.book[sheetname].column_dimensions['AF'].width = 50

    # Formatiere Datum
    writer.book[sheetname].cell(row=1, column=1, value="Erstellt am " + date).style = format_title
    # Formatiere Streckenname
    writer.book[sheetname].cell(row=2, column=1, value=name).style = format_title
    # Suche letzte Zeile
    row_line = 3
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=2, max_col=2, max_row=4000):
        for cell in row:
            if not cell.value is None:
                row_line += 1
    # Suche letzte Spalte
    column = 1
    for cols in writer.book[sheetname].iter_cols(min_row=4, min_col=1, max_col=999, max_row=4):
        for cell in cols:
            if not cell.value is None:
                column += 1
    # Formatiere Header linksbündig für alle weiteren Spalten, da sonst zu kompliziert im Moment
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=1, max_col=column, max_row=4):
        for cell in row:
            cell.style = format_header_left
    # Formatiere iRacingNr
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=1, max_col=1, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Formatiere Name
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=2, max_col=2, max_row=row_line):
        for cell in row:
            cell.style = format_driver
    # Formatiere iRacingName
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=3, max_col=3, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Formatiere Serie, Team, Date, Track
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=4, max_col=column, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Weisser Hintergrund
    for row in writer.book[sheetname].iter_rows(min_row=1, min_col=2, max_col=column, max_row=2):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=3, min_col=1, max_col=column, max_row=3):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=row_line + 1, min_col=2, max_col=column, max_row=row_line + 1):
        for cell in row:
            cell.style = format_background
    writer.save()


def format_driver_export(payload):
    dataframe = payload[0]
    date = payload[1]
    name = payload[2]
    sheetname_long = name
    sheetname = sheetname_long[:30]
    if 'Sheet' in writer.book.sheetnames:
        del writer.book['Sheet']
    dataframe.to_excel(writer, sheet_name=sheetname, startrow=3, index=False)
    writer.book[sheetname].column_dimensions['A'].width = 20
    writer.book[sheetname].column_dimensions['B'].width = 30
    writer.book[sheetname].column_dimensions['C'].width = 40
    writer.book[sheetname].column_dimensions['D'].width = 20
    writer.book[sheetname].column_dimensions['E'].width = 50
    writer.book[sheetname].column_dimensions['F'].width = 20
    writer.book[sheetname].column_dimensions['G'].width = 50
    writer.book[sheetname].column_dimensions['H'].width = 10
    writer.book[sheetname].column_dimensions['I'].width = 10
    writer.book[sheetname].column_dimensions['J'].width = 10
    writer.book[sheetname].column_dimensions['K'].width = 10
    # Formatiere Datum
    writer.book[sheetname].cell(row=1, column=1, value="Erstellt am " + date).style = format_title
    # Formatiere Streckenname
    writer.book[sheetname].cell(row=2, column=1, value="Für Fahrer*in: " + name).style = format_title
    # Formatiere Header zentriert für FlancNr
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=1, max_col=1, max_row=4):
        for cell in row:
            cell.style = format_header_center
    # Formatiere Header linksbündig für Driver, Name, Serie, Team, Date, Track... Zum Glück :9
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=2, max_col=7, max_row=4):
        for cell in row:
            cell.style = format_header_left
    # Formatiere Header entriert für Pos, Starts, Incs und Points
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=8, max_col=11, max_row=4):
        for cell in row:
            cell.style = format_header_center
    # Suche letzte Zeile
    row_line = 3
    for row in writer.book[sheetname].iter_rows(min_row=4, min_col=2, max_col=2, max_row=4000):
        for cell in row:
            if not cell.value is None:
                row_line += 1
    # Formatiere FlancNr
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=1, max_col=1, max_row=row_line):
        for cell in row:
            cell.style = format_pos
    # Formatiere Driver
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=2, max_col=2, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Formatiere Name
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=3, max_col=3, max_row=row_line):
        for cell in row:
            cell.style = format_driver
    # Formatiere Serie, Team, Date, Track
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=4, max_col=7, max_row=row_line):
        for cell in row:
            cell.style = format_team
    # Formatiere Pos, Starts und Incs, Punkte
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=8, max_col=10, max_row=row_line):
        for cell in row:
            cell.style = format_start_inc
    # Formatiere Points
    for row in writer.book[sheetname].iter_rows(min_row=5, min_col=11, max_col=11, max_row=row_line):
        for cell in row:
            cell.style = format_pos
    # Weisser Hintergrund
    for row in writer.book[sheetname].iter_rows(min_row=1, min_col=2, max_col=11, max_row=2):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=3, min_col=1, max_col=11, max_row=3):
        for cell in row:
            cell.style = format_background
    for row in writer.book[sheetname].iter_rows(min_row=row_line + 1, min_col=2, max_col=11, max_row=row_line + 1):
        for cell in row:
            cell.style = format_background
    writer.save()
